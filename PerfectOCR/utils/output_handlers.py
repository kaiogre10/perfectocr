# PerfectOCR/utils/output_handlers.py
import json
import os
import logging
from typing import Dict, Any, Optional, List
from .encoders import NumpyEncoder # Asumiendo que NumpyEncoder está en utils.encoders
from utils.table_formatter import TableFormatter

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl no está disponible. ExcelOutputHandler no funcionará.")

logger = logging.getLogger(__name__)

class JsonOutputHandler:
    """
    Obrero especializado en guardar datos en formato JSON.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        self.enabled_outputs = self.config.get('enabled_outputs', {})
        
    def should_save_output(self, output_type: str) -> bool:
        """
        Verifica si un tipo específico de output está habilitado.
        Si un tipo de output no está en la configuración, se asume que está deshabilitado.
        """
        return self.enabled_outputs.get(output_type, False)
        
    def save(self, data: Dict[str, Any], output_dir: str, file_name_with_extension: str, output_type: str = None) -> Optional[str]:
        """
        Guarda un diccionario de datos en un archivo JSON si el tipo de output está habilitado.
        
        Args:
            data: El diccionario a guardar.
            output_dir: El directorio donde se guardará el archivo.
            file_name_with_extension: El nombre del archivo.
            output_type: El tipo de output (e.g., 'ocr_raw', 'reconstructed_lines', etc.)
        """
        if output_type and not self.should_save_output(output_type):
            logger.debug(f"Output {output_type} está deshabilitado, omitiendo guardado.")
            return None
            
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Datos JSON guardados en: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error guardando JSON en {output_path}: {e}", exc_info=True)
            return None
            
class TextOutputHandler:
    """
    Obrero especializado en guardar contenido de texto plano.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        # logger.debug("TextOutputHandler initialized.")

    def save(self, text_content: str, output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda una cadena de texto en un archivo.
        Crea el directorio de salida si no existe.

        Args:
            text_content: La cadena de texto a guardar.
            output_dir: El directorio donde se guardará el archivo.
            file_name_with_extension: El nombre del archivo (e.g., "transcription.txt").

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(text_content)
            logger.info(f"Datos de texto guardados en: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error guardando archivo de texto en {output_path}: {e}", exc_info=True)
            return None
        
class MarkdownOutputHandler:
    """
    Obrero especializado en generar y guardar contenido Markdown,
    especialmente tablas.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        # logger.debug("MarkdownOutputHandler initialized.")

    def save_table_view(self, 
                        headers: List[str], 
                        table_matrix: List[List[Dict[str, Any]]], 
                        output_dir: str, 
                        file_name_with_extension: str,
                        document_title: Optional[str] = None) -> Optional[str]:
        """
        Formatea una tabla como Markdown y la guarda en un archivo.

        Args:
            headers: Lista de encabezados de columna.
            table_matrix: Matriz de la tabla (lista de filas, donde cada fila es lista de celdas-dict).
            output_dir: Directorio de salida.
            file_name_with_extension: Nombre del archivo Markdown.
            document_title: Título opcional para el documento Markdown.

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        markdown_content_lines = []
        if document_title:
            markdown_content_lines.append(f"# Tabla Extraída para: {document_title}\n")

        markdown_table_str = TableFormatter.format_as_markdown(headers, table_matrix) #
        markdown_content_lines.append(markdown_table_str)

        full_markdown_content = "\n".join(markdown_content_lines)

        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f_md:
                f_md.write(full_markdown_content)
            logger.info(f"Vista de tabla Markdown guardada en: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error guardando Markdown de tabla en {output_path}: {e}", exc_info=True)
            return None

class ExcelOutputHandler:
    """
    Obrero especializado en generar y guardar archivos Excel,
    especialmente para matrices semánticamente corregidas.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        if not OPENPYXL_AVAILABLE:
            logger.error("ExcelOutputHandler requiere openpyxl. Instala con: pip install openpyxl")
        logger.debug("ExcelOutputHandler inicializado.")

    def save_semantically_corrected_matrices(self, 
                                           matrices_data: List[Dict[str, Any]], 
                                           output_dir: str, 
                                           file_name: str = "ground_truth_batch.xlsx") -> Optional[str]:
        """
        Guarda múltiples matrices semánticamente corregidas en un archivo Excel.
        Mantiene exactamente la misma estructura que el JSON original.

        Args:
            matrices_data: Lista de diccionarios con estructura:
                          {
                              "document_id": str,
                              "headers": List[str],
                              "semantic_types": List[str], 
                              "matrix": List[List[str]],
                              "error": str (opcional)
                          }
            output_dir: Directorio de salida.
            file_name: Nombre del archivo Excel.

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("No se puede guardar Excel: openpyxl no está disponible")
            return None
            
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ground Truth Matrices"
            
            # Preparar datos para Excel
            all_rows = []
            
            # Agregar encabezado principal
            if matrices_data:
                first_matrix = next((m for m in matrices_data if "headers" in m), None)
                if first_matrix:
                    headers = ["ID_DOCUMENTO", "ESTADO", "ERROR"] + first_matrix.get("headers", [])
                    semantic_types = ["metadata", "metadata", "metadata"] + first_matrix.get("semantic_types", [])
                    
                    # Primera fila: Headers
                    all_rows.append(headers)
                    # Segunda fila: Semantic Types
                    all_rows.append(semantic_types)
                    
                    # Agregar datos de todas las matrices
                    for matrix_data in matrices_data:
                        document_id = matrix_data.get("document_id", "unknown")
                        error = matrix_data.get("error", "")
                        status = "ERROR" if error else "OK"
                        
                        if "matrix" in matrix_data:
                            matrix = matrix_data.get("matrix", [])
                            for row in matrix:
                                excel_row = [document_id, status, error] + row
                                all_rows.append(excel_row)
                        else:
                            # Si no hay matriz, agregar una fila indicando el error
                            excel_row = [document_id, status, error] + [""] * len(first_matrix.get("headers", []))
                            all_rows.append(excel_row)
            
            # Escribir todas las filas al worksheet
            for row in all_rows:
                ws.append(row)
            
            # Guardar archivo
            wb.save(output_path)
            logger.info(f"Archivo Excel de ground truth guardado en: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error guardando archivo Excel en {output_path}: {e}", exc_info=True)
            return None

    def append_matrix_to_existing_excel(self, 
                                      matrix_data: Dict[str, Any], 
                                      excel_path: str) -> bool:
        """
        Añade una nueva matriz a un archivo Excel existente.

        Args:
            matrix_data: Diccionario con estructura:
                        {
                            "document_id": str,
                            "headers": List[str],
                            "semantic_types": List[str], 
                            "matrix": List[List[str]],
                            "error": str (opcional)
                        }
            excel_path: Ruta al archivo Excel existente.

        Returns:
            True si se añadió correctamente, False en caso contrario.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("No se puede modificar Excel: openpyxl no está disponible")
            return False
            
        try:
            if not os.path.exists(excel_path):
                logger.warning(f"Archivo Excel no existe: {excel_path}")
                return False
                
            # Cargar workbook existente
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Obtener datos de la nueva matriz
            document_id = matrix_data.get("document_id", "unknown")
            error = matrix_data.get("error", "")
            status = "ERROR" if error else "OK"
            matrix = matrix_data.get("matrix", [])
            
            # Añadir filas de la nueva matriz con el formato correcto
            if matrix:
                for row in matrix:
                    excel_row = [document_id, status, error] + row
                    ws.append(excel_row)
            else:
                # Si no hay matriz, agregar una fila indicando el error
                # Necesitamos saber cuántas columnas tiene el Excel existente
                if ws.max_column > 3:  # ID_DOCUMENTO, ESTADO, ERROR + headers
                    num_data_columns = ws.max_column - 3
                    excel_row = [document_id, status, error] + [""] * num_data_columns
                    ws.append(excel_row)
            
            # Guardar archivo
            wb.save(excel_path)
            logger.info(f"Matriz añadida al archivo Excel: {excel_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error añadiendo matriz al archivo Excel {excel_path}: {e}", exc_info=True)
            return False

#futuros módulos