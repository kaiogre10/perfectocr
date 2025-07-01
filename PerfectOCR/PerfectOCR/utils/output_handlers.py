# PerfectOCR/utils/output_handlers.py
import json
import os
import logging
from typing import Dict, Any, Optional, List
from .encoders import NumpyEncoder # Asumiendo que NumpyEncoder está en utils.encoders
from utils.table_formatter import TableFormatter

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl no está disponible. ExcelOutputHandler no funcionará.")

logger = logging.getLogger(__name__)

class JsonOutputHandler:
    """
    Obrero especializado en guardar datos en formato JSON.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config or {}

    def should_save_output(self, output_type: str) -> bool:
        enabled_outputs = self.config.get("enabled_outputs", {})
        return enabled_outputs.get(output_type, False)
        
    def save(self, data: Dict[str, Any], output_dir: str, file_name_with_extension: str, output_type: str = None) -> Optional[str]:
        """
        Guarda datos en formato JSON en un archivo.
        Crea el directorio de salida si no existe.

        Args:
            data: Los datos a guardar en formato JSON.
            output_dir: El directorio donde se guardará el archivo.
            file_name_with_extension: El nombre del archivo.
            output_type: El tipo de output (e.g., 'ocr_raw', 'reconstructed_lines', etc.)
        """
        if output_type and not self.should_save_output(output_type):
            logger.debug(f"Output {output_type} está deshabilitado, omitiendo guardado.")
            return None
            
        # Validar que output_dir no esté vacío
        if not output_dir or output_dir.strip() == "":
            logger.error("Error: output_dir está vacío o es None")
            return None
            
        output_path = None
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Datos JSON guardados en: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Error guardando JSON"
            if output_path:
                error_msg += f" en {output_path}"
            error_msg += f": {e}"
            logger.error(error_msg, exc_info=True)
            return None
            
    def save_reconstructed_lines_text_view(self, reconstructed_data: Dict[str, Any], output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda una vista de depuración de las líneas reconstruidas, mostrando solo el texto de cada línea por motor.
        """
        text_view_data = {}
        # Unificar datos de múltiples documentos si es una lista
        if isinstance(reconstructed_data, list) and all("document_id" in item for item in reconstructed_data):
            # Procesa una lista de resultados de documentos
            for doc_data in reconstructed_data:
                doc_id = doc_data.get("document_id", "unknown_doc")
                doc_text_lines = {}
                for engine_name, lines in doc_data.items():
                    if engine_name == 'page_dimensions' or engine_name == 'document_id' or not isinstance(lines, list):
                        continue
                    doc_text_lines[engine_name] = [line.get('text_raw', '') for line in lines]
                text_view_data[doc_id] = doc_text_lines
        else: # Procesa un solo documento
            for engine_name, lines in reconstructed_data.items():
                if engine_name == 'page_dimensions' or not isinstance(lines, list):
                    continue
                text_view_data[f"{engine_name}_reconstructed_text"] = [line.get('text_raw', '') for line in lines]

        if not text_view_data:
            logger.info("No hay datos de líneas reconstruidas para generar la vista de texto de depuración.")
            return None

        # Llamar al método 'save' genérico para escribir el archivo JSON.
        # NOTA: el flag 'should_save_output' se controla en el coordinador que llama a este método.
        output_path = None
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(text_view_data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Vista de texto de depuración guardada en: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Error guardando la vista de texto de depuración"
            if output_path:
                error_msg += f" en {output_path}"
            error_msg += f": {e}"
            logger.error(error_msg, exc_info=True)
            return None

    def save_cleaned_lines_text_view(self, cleaned_data: Dict[str, Any], output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda una vista de depuración de las líneas limpias, mostrando solo el texto de cada línea por motor.
        """
        text_view_data = {}
        for engine_name, lines in cleaned_data.items():
            if engine_name == 'page_dimensions' or not isinstance(lines, list):
                continue
            
            text_view_data[f"{engine_name}_cleaned_text"] = [line.get('text_raw', '') for line in lines]

        if not text_view_data:
            logger.info("No hay datos de líneas limpias para generar la vista de texto de depuración.")
            return None

        # Llamar al método 'save' genérico para escribir el archivo JSON.
        output_path = None
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(text_view_data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Vista de texto de depuración (cleaned) guardada en: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Error guardando la vista de texto de depuración (cleaned)"
            if output_path:
                error_msg += f" en {output_path}"
            error_msg += f": {e}"
            logger.error(error_msg, exc_info=True)
            return None

    def save_corrected_lines_text_view(self, corrected_data: Dict[str, Any], output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda una vista de depuración de las líneas corregidas, mostrando solo el texto de cada línea por motor.
        """
        text_view_data = {}
        for engine_name, lines in corrected_data.items():
            if engine_name == 'page_dimensions' or not isinstance(lines, list):
                continue
            
            text_view_data[f"{engine_name}_corrected_text"] = [line.get('text_raw', '') for line in lines]

        if not text_view_data:
            logger.info("No hay datos de líneas corregidas para generar la vista de texto de depuración.")
            return None

        # Llamar al método 'save' genérico para escribir el archivo JSON.
        output_path = None
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(text_view_data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Vista de texto de depuración (corrected) guardada en: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Error guardando la vista de texto de depuración (corrected)"
            if output_path:
                error_msg += f" en {output_path}"
            error_msg += f": {e}"
            logger.error(error_msg, exc_info=True)
            return None

    def save_geo_matrix_consolidated(self, consolidated_data: Dict[str, Any], output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda el output consolidado completo del geo_matrix con toda la metadata.
        Incluye matriz, headers, totales, dimensiones y información de procesamiento.
        """
        if not consolidated_data:
            logger.warning("No hay datos consolidados del geo_matrix para guardar.")
            return None

        output_path = None
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(consolidated_data, f, indent=4, ensure_ascii=False, cls=NumpyEncoder)
            logger.info(f"Output consolidado del geo_matrix guardado en: {output_path}")
            return output_path
        except Exception as e:
            error_msg = f"Error guardando output consolidado del geo_matrix"
            if output_path:
                error_msg += f" en {output_path}"
            error_msg += f": {e}"
            logger.error(error_msg, exc_info=True)
            return None

class TextOutputHandler:
    """
    Obrero especializado en guardar contenido de texto plano.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        # logger.debug("TextOutputHandler initialized.")

    def save(self, text_content: str, output_dir: str, file_name_with_extension: str) -> Optional[str]:
        """
        Guarda una cadena de texto en un archivo.
        Crea el directorio de salida si no existe.

        Args:
            text_content: La cadena de texto a guardar.
            output_dir: El directorio donde se guardará el archivo.
            file_name_with_extension: El nombre del archivo (e.g., "transcription.txt").

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(text_content)
            logger.info(f"Datos de texto guardados en: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error guardando archivo de texto en {output_path}: {e}", exc_info=True)
            return None
        
class MarkdownOutputHandler:
    """
    Obrero especializado en generar y guardar contenido Markdown,
    especialmente tablas.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        # logger.debug("MarkdownOutputHandler initialized.")

    def save_table_view(self, 
                        headers: List[str], 
                        table_matrix: List[List[Dict[str, Any]]], 
                        output_dir: str, 
                        file_name_with_extension: str,
                        document_title: Optional[str] = None) -> Optional[str]:
        """
        Formatea una tabla como Markdown y la guarda en un archivo.

        Args:
            headers: Lista de encabezados de columna.
            table_matrix: Matriz de la tabla (lista de filas, donde cada fila es lista de celdas-dict).
            output_dir: Directorio de salida.
            file_name_with_extension: Nombre del archivo Markdown.
            document_title: Título opcional para el documento Markdown.

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        markdown_content_lines = []
        if document_title:
            markdown_content_lines.append(f"# Tabla Extraída para: {document_title}\n")

        markdown_table_str = TableFormatter.format_as_markdown(headers, table_matrix) #
        markdown_content_lines.append(markdown_table_str)

        full_markdown_content = "\n".join(markdown_content_lines)

        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name_with_extension)
            with open(output_path, 'w', encoding='utf-8') as f_md:
                f_md.write(full_markdown_content)
            logger.info(f"Vista de tabla Markdown guardada en: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error guardando Markdown de tabla en {output_path}: {e}", exc_info=True)
            return None

class ExcelOutputHandler:
    """
    Obrero especializado en generar y guardar archivos Excel,
    especialmente para matrices con resultados finales.
    """
    def __init__(self, config: Optional[Dict] = None):
        self.config = config if config is not None else {}
        if not OPENPYXL_AVAILABLE:
            logger.error("ExcelOutputHandler requiere openpyxl. Instala con: pip install openpyxl")
        logger.debug("ExcelOutputHandler inicializado.")

    def save_math_max_matrices(self, 
                              matrices_data: List[Dict[str, Any]], 
                              output_dir: str, 
                              file_name: str = "math_max_resultados_finales.xlsx") -> Optional[str]:
        """
        Guarda múltiples matrices math_max_matrix en un archivo Excel.
        Diseñado específicamente para los resultados finales de math_max.

        Args:
            matrices_data: Lista de diccionarios con estructura de math_max_matrix.
            output_dir: Directorio de salida.
            file_name: Nombre del archivo Excel.

        Returns:
            La ruta completa al archivo guardado si tiene éxito, None en caso contrario.
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("No se puede guardar Excel: openpyxl no está disponible")
            return None
            
        output_path = ""
        try:
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, file_name)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados Finales"
            
            all_rows = []
            
            if matrices_data:
                first_valid_matrix = next((m for m in matrices_data if m.get("headers") and m.get("matrix")), None)
                
                if first_valid_matrix:
                    excel_headers = ["ID_DOCUMENTO", "ESTADO", "ERROR"] + first_valid_matrix.get("headers", [])
                    semantic_types = ["metadata", "metadata", "metadata"] + first_valid_matrix.get("semantic_types", [])
                    
                    all_rows.append(excel_headers)
                    all_rows.append(semantic_types)
                    
                    for matrix_data in matrices_data:
                        document_id = matrix_data.get("document_id", "unknown")
                        error = matrix_data.get("error", "")
                        semantic_format = matrix_data.get("semantic_format", "")
                        
                        status = "ERROR" if error else ("MATH_CORRECTED" if semantic_format == "math_assigned" else "PROCESSED")
                        
                        if matrix_data.get("matrix"):
                            for row in matrix_data["matrix"]:
                                excel_row = [document_id, status, error] + row
                                all_rows.append(excel_row)
                        else:
                            excel_row = [document_id, status, error] + [""] * len(first_valid_matrix.get("headers", []))
                            all_rows.append(excel_row)
                        
                        all_rows.append([""] * len(excel_headers)) # Separador
            
            for row in all_rows:
                ws.append(row)
            
            if all_rows:
                for cell in ws[1]: cell.font = cell.font.copy(bold=True)
                for cell in ws[2]: cell.font = cell.font.copy(italic=True)
            
            for column in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(output_path)
            logger.info(f"Archivo Excel de resultados finales guardado en: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error guardando archivo Excel de resultados finales en {output_path}: {e}", exc_info=True)
            return None

#futuros módulos